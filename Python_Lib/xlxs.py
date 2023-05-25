from openpyxl import load_workbook
import datetime

# def xlxs_data(runcard,modbus):
class excel:
    def __init__(self,sheet):

        self.data_A = []
        self.data_B = []
        self.data_C = []
        self.sheet = sheet
        self.xlxs_read()
        try:
            for i in range(len(self.data_A)):
                        self.date.append("'"+self.data_A[i]+" "+ self.data_B[i])
        except:pass
    
    def xlxs_read(self):
        wb = load_workbook(filename = r'C:\Users\filmm\OneDrive\Desktop\UntitledProject\excel\xcel.xlsx')
        self.worksheet = wb[self.sheet]
        A = self.worksheet['a']
        B = self.worksheet['b']
        C = self.worksheet['c']
        self.data_A.clear()
        self.data_B.clear()
        self.data_C.clear()

        x =-1
        for i in A:
            x+=1
            if str(i.value) != "None":
                self.data_A.append(str(i.value))

        x =-1
        for i in B:
            x+=1
            if str(i.value) != "None":
                self.data_B.append(str(i.value))

        x =-1
        for i in C:
            x+=1
            if str(i.value) != "None":
                self.data_C.append(str(i.value))
        
        # self.data_A.reverse()
        # self.data_B.reverse()
        # self.data_C.reverse()
        del self.data_A[0]
        del self.data_B[0]
        del self.data_C[0]
        wb.close()

        return self.data_A,self.data_B,self.data_C
    date = []
    def xlxs_write(self,runcard):
        date = datetime.datetime.now()
        
        if len(runcard) >5: 
            if(runcard in self.data_C):
                # modbus.write(0,0)
                # messagebox("Runcard duplicate")
                return self.data_C,self.date,0
            
            else:
                wb = load_workbook(filename = r'C:\Users\filmm\OneDrive\Desktop\UntitledProject\excel\xcel.xlsx')
                self.worksheet = wb[self.sheet]
                self.worksheet.append({'A' : str(date.strftime("%d-%m-%Y")), 'B' : str(date.strftime("%X")), 'C' : f'{runcard}'})
                wb.save(r'C:\Users\filmm\OneDrive\Desktop\UntitledProject\excel\xcel.xlsx')
                wb.close()
                self.xlxs_read()
                print(self.data_C)
                for i in range(len(self.data_A)-1):
                        self.date.append("'"+self.data_A[i]+" "+ self.data_B[i])
                    
                return self.data_C,self.date,1
        return self.data_C,self.date,0

if __name__ == "__main__":
    x = excel("accept")
    for i in range(10):
        print(x.xlxs_write(f"th0000-{i}"))
