from openpyxl import load_workbook
import datetime

# def xlxs_data(runcard,modbus):
class excel:
    def __init__(self):

        self.data_A = []
        self.data_B = []
        self.data_C = []
        self.reject_A = []
        self.reject_B = []
        self.reject_C = []
    
        self.sheet = "font_line"
        self.xlxs_read()
        self.reject_read()
        try:
            for i in range(len(self.reject_A)):
                            self.date.append("'"+self.reject_A[i]+" "+ self.reject_B[i])
        except:pass
    
    def xlxs_read(self):
        wb = load_workbook(filename = r'C:\Users\filmm\OneDrive\Desktop\UntitledProject\excel\xcel.xlsx')
        self.worksheet = wb[self.sheet]
        A = self.worksheet['a']
        B = self.worksheet['b']
        C = self.worksheet['c']
        self.data_A.clear()
        self.data_B.clear()
        self.data_C.clear()

        x =-1
        for i in A:
            x+=1
            if str(i.value) != "None":
                self.data_A.append(str(i.value))

        x =-1
        for i in B:
            x+=1
            if str(i.value) != "None":
                self.data_B.append(str(i.value))

        x =-1
        for i in C:
            x+=1
            if str(i.value) != "None":
                self.data_C.append(str(i.value))
        
        # self.data_A.reverse()
        # self.data_B.reverse()
        # self.data_C.reverse()
        del self.data_A[0]
        del self.data_B[0]
        del self.data_C[0]
        wb.close()

        return self.data_A,self.data_B,self.data_C
    
    def reject_read(self):
        wb = load_workbook(filename = r'C:\Users\filmm\OneDrive\Desktop\UntitledProject\excel\xcel.xlsx')
        self.worksheet = wb["reject"]
        A = self.worksheet['a']
        B = self.worksheet['b']
        C = self.worksheet['c']
        self.reject_A.clear()
        self.reject_B.clear()
        self.reject_C.clear()

        x =-1
        for i in A:
            x+=1
            if str(i.value) != "None":
                self.reject_A.append(str(i.value))

        x =-1
        for i in B:
            x+=1
            if str(i.value) != "None":
                self.reject_B.append(str(i.value))

        x =-1
        for i in C:
            x+=1
            if str(i.value) != "None":
                self.reject_C.append(str(i.value))
        
        # self.data_A.reverse()
        # self.data_B.reverse()
        # self.data_C.reverse()
        del self.reject_A[0]
        del self.reject_B[0]
        del self.reject_C[0]
        wb.close()

        return self.reject_A,self.reject_B,self.reject_C
        
    
         
    date = []
    def xlxs_write(self,runcard):
        date = datetime.datetime.now()
        
        if len(runcard) >5: 
            if runcard in self.data_C:
                dddd = runcard in self.reject_C
                if dddd == False:
                    wb = load_workbook(filename = r'C:\Users\filmm\OneDrive\Desktop\UntitledProject\excel\xcel.xlsx')
                    self.worksheet = wb["reject"]
                    self.worksheet.append({'A' : str(date.strftime("%d-%m-%Y")), 'B' : str(date.strftime("%X")), 'C' : f'{runcard}'})
                    wb.save(r'C:\Users\filmm\OneDrive\Desktop\UntitledProject\excel\xcel.xlsx')
                    wb.close()
                    self.xlxs_read()
                    self.reject_read()
                    # print(self.data_C)
                    for i in range(len(self.reject_A)):
                            self.date.append("'"+self.reject_A[i]+" "+ self.reject_B[i])
                            
                    return self.reject_C,self.date,"reject",1
                else: return self.reject_C,self.date,"reject",0
            else:
                return self.reject_C,self.date,"accept",0
        return self.reject_C,self.date,"accept",0

if __name__ == "__main__":
    x = excel()
    for i in range(10):
        d = x.xlxs_write(f"4987072024201")
        print(len(d [1]))
        
